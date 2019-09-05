from openpyxl import load_workbook
from collections import namedtuple

class ReadXls:
    def __init__(self,filename):
        #data_only=True代表連formula也是計算後的結果值
        self.wb = load_workbook(filename = filename,data_only=True) 

    def fieldTitle(self,sheetName,rowPoisition):
        """取得欄位名稱"""
        ws = self.wb[sheetName]
        rowObj = []
        for row in ws.iter_rows(min_row=rowPoisition, max_col=ws.max_column, max_row=rowPoisition): 
            rowObj = ([row[i].value for i in range(ws.max_column) if row[i].value != None])
        return rowObj

    def getSheetData(self,sheetName,startRow = 3,endcol = 0):
        """取得指定資料表的內容"""
        ws = self.wb[sheetName]
        rows = []
        for row in ws.iter_rows(min_row=startRow, max_col=endcol, max_row=ws.max_row): 
            if row[0].value == None:
                print('this row is None')
            else:
                rows.append([row[i] for i in range(endcol)])
        return rows
        # for row in rows: 
        #     for cell in row:
        #         print(cell.value)

    def getSheetNames(self,NotIn = [] ):
        """取得工作表名稱"""
        return [sheetName for sheetName in self.wb.sheetnames if sheetName not in NotIn]

    def GetRows(self,sheetName,titles,startRow = 4):    
        colLen = len(titles)
        rows = self.getSheetData(sheetName,startRow,colLen)
        mapRows = []
        for row in rows: 
            Fields=namedtuple("Fields",titles)
            myList = [field.value for field in row]
            fields = Fields(*myList)
            mapRows.append(fields)
        return mapRows
